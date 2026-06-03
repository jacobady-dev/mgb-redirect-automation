#!/usr/bin/env python3
"""Local AI review pass for MGB redirect workbooks.

This script is intentionally local-only. It reads a generated redirect workbook,
reviews rows that are likely to need judgment, and writes AI review columns back
into a copy of the workbook.

Default behavior is non-destructive:
- It does not overwrite Destination URL.
- It does not overwrite DW - Approval.
- It adds adjacent AI review columns.

Set OPENAI_API_KEY in your local environment before running.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from openai import OpenAI
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "The OpenAI Python package is required. Run: pip install -r requirements.txt"
    ) from exc


DEFAULT_MODEL = os.getenv("OPENAI_MODEL", "gpt-5.5-mini")
AI_COLUMNS = [
    "AI Decision",
    "AI Risk Level",
    "AI Recommended Destination",
    "AI Notes",
    "AI Flags",
]
REVIEWABLE_SHEETS = ["HUMAN CHECK", "Redirect List for CSV"]
DESTINATION_SHEET = "MGB Destination Candidates"

STOP_WORDS = {
    "https",
    "http",
    "www",
    "com",
    "org",
    "en",
    "html",
    "aspx",
    "mass",
    "general",
    "brigham",
    "newton",
    "wellesley",
    "hospital",
    "hospitals",
    "page",
    "pages",
    "about",
    "care",
    "service",
    "services",
    "patient",
    "patients",
    "visitor",
    "visitors",
    "medical",
    "center",
    "centers",
    "and",
    "the",
    "for",
    "with",
    "from",
}


@dataclass
class Candidate:
    url: str
    title: str = ""
    h1: str = ""
    destination_type: str = "unknown"
    score: float = 0.0


@dataclass
class ReviewTarget:
    sheet_name: str
    row_number: int
    existing_url: str
    current_destination: str
    dw_approval: str
    notes: str
    content_type: str
    status_code: str
    title: str
    source_type: str
    flags: List[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="AI review generated MGB redirect workbook rows.")
    parser.add_argument("workbook", type=Path, help="Generated redirect workbook .xlsx")
    parser.add_argument(
        "--output",
        type=Path,
        help="Output workbook path. Defaults to *_ai_reviewed.xlsx",
    )
    parser.add_argument("--model", default=DEFAULT_MODEL, help="OpenAI model to use")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=100,
        help="Maximum rows to review in this run. Use 0 for no limit.",
    )
    parser.add_argument(
        "--include-clean-redirects",
        action="store_true",
        help="Also review Redirect List rows without REVIEW/risk flags. Off by default.",
    )
    parser.add_argument(
        "--apply-ai-destination",
        action="store_true",
        help="Overwrite Destination URL with AI recommendation when AI decision is approve/review. Off by default.",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.2,
        help="Delay between API calls in seconds.",
    )
    return parser.parse_args()


def norm(value: Any) -> str:
    return str(value or "").strip()


def get_headers(ws: Worksheet) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for cell in ws[1]:
        value = norm(cell.value)
        if value:
            headers[value] = cell.column
    return headers


def ensure_columns(ws: Worksheet, columns: Sequence[str]) -> Dict[str, int]:
    headers = get_headers(ws)
    next_col = ws.max_column + 1
    for column_name in columns:
        if column_name not in headers:
            ws.cell(row=1, column=next_col, value=column_name)
            headers[column_name] = next_col
            next_col += 1
    return headers


def cell_value(ws: Worksheet, row: int, headers: Dict[str, int], *names: str) -> str:
    for name in names:
        col = headers.get(name)
        if col:
            value = norm(ws.cell(row=row, column=col).value)
            if value:
                return value
    return ""


def write_cell(ws: Worksheet, row: int, headers: Dict[str, int], name: str, value: str) -> None:
    col = headers.get(name)
    if col:
        ws.cell(row=row, column=col, value=value)


def tokens(text: str) -> set[str]:
    raw = re.sub(r"[^a-z0-9]+", " ", text.lower()).split()
    return {token.rstrip("s") for token in raw if len(token) > 2 and token not in STOP_WORDS}


def destination_type(url: str, title: str = "") -> str:
    text = f"{url} {title}".lower()
    if "doctors.massgeneralbrigham.org" in text or "/providers" in text:
        return "provider"
    if "/locations/" in text or "/find-a-location" in text:
        return "location"
    if "/about/newsroom/" in text or "/press-releases/" in text:
        return "news_article"
    if "/patient-stories/" in text or "/stories" in text:
        return "patient_story"
    if "/about/careers" in text:
        return "careers"
    if "/billing" in text or "medical-records" in text or "patient-gateway" in text:
        return "admin_resource"
    if "/patient-care/services-and-specialties" in text:
        return "service"
    if "/patient-care/patient-visitor-information" in text:
        return "patient_visitor"
    if "/about/" in text:
        return "about"
    return "unknown"


def source_type(source_url: str, title: str = "", notes: str = "") -> str:
    text = f"{source_url} {title} {notes}".lower()
    if any(x in text for x in [".pdf", ".doc", ".zip", ".xml", ".jpg", ".png", ".js", ".css"]):
        return "asset"
    if "find-a-doctor" in text or "provider" in text or "doctor profile" in text:
        return "provider"
    if "/news" in text or "press release" in text or "article" in text or "community update" in text:
        return "news_article"
    if "patient-stories" in text or "patient story" in text:
        return "patient_story"
    if "/locations" in text or "directions" in text or "parking" in text:
        return "location"
    if "careers" in text or "employment" in text or "job" in text:
        return "careers"
    if any(x in text for x in ["billing", "insurance", "medical records", "patient gateway", "charge data"]):
        return "admin_resource"
    if any(
        x in text
        for x in [
            "/medical-services/",
            "/orthopedics/",
            "/radiology/",
            "/maternity/",
            "/rehabilitation-services/",
            "/mass-general-cancer-center/",
            "diabetes",
            "endocrinology",
            "cardiology",
            "cancer",
            "orthopedics",
            "rehabilitation",
        ]
    ):
        return "service"
    if "patients-and-visitors" in text or "classes-and-resources" in text:
        return "patient_visitor"
    return "unknown"


def risk_flags(source: ReviewTarget, destination_url: str) -> List[str]:
    flags: List[str] = []
    dest_type = destination_type(destination_url)
    url = destination_url.lower()

    if not destination_url:
        flags.append("missing_destination")
    if destination_url and "massgeneralbrigham.org" not in url:
        flags.append("external_destination")
    if source.source_type == "service" and dest_type == "location":
        flags.append("service_to_location")
    if source.source_type == "service" and dest_type in {"news_article", "patient_story"}:
        flags.append("service_to_content_page")
    if source.source_type in {"news_article", "patient_story"} and dest_type not in {"news_article", "patient_story"}:
        flags.append("content_page_without_specific_content_match")
    if source.source_type == "careers" and "fraud-alert" in url:
        flags.append("careers_to_fraud_alert")
    if source.source_type != "location" and "/locations/" in url:
        flags.append("non_location_to_location")
    if source.source_type != "news_article" and "/about/newsroom/" in url:
        flags.append("non_news_to_newsroom")
    if any(host in url for host in ["salem.massgeneralbrigham.org", "cpdlearn.massgeneralbrigham.org", "cooleydickinson"]):
        flags.append("other_site_or_learning_domain")
    return sorted(set(flags))


def read_candidates(wb) -> List[Candidate]:
    if DESTINATION_SHEET not in wb.sheetnames:
        return []
    ws = wb[DESTINATION_SHEET]
    headers = get_headers(ws)
    candidates: List[Candidate] = []
    for row in range(2, ws.max_row + 1):
        url = cell_value(ws, row, headers, "Destination URL", "Address", "URL")
        if not url:
            continue
        title = cell_value(ws, row, headers, "Title", "Title 1", "H1", "H1-1")
        if "massgeneralbrigham.org" not in url:
            continue
        candidates.append(Candidate(url=url, title=title, h1=title, destination_type=destination_type(url, title)))
    return candidates


def top_candidates(target: ReviewTarget, candidates: Sequence[Candidate], limit: int = 8) -> List[Candidate]:
    source_tokens = tokens(f"{target.existing_url} {target.title} {target.notes} {target.source_type}")
    scored: List[Candidate] = []
    for candidate in candidates:
        candidate_tokens = tokens(f"{candidate.url} {candidate.title} {candidate.h1} {candidate.destination_type}")
        overlap = len(source_tokens & candidate_tokens)
        type_bonus = 2 if candidate.destination_type == target.source_type else 0
        service_bonus = 1 if target.source_type == "service" and candidate.destination_type == "service" else 0
        candidate.score = overlap + type_bonus + service_bonus
        if candidate.score > 0:
            scored.append(candidate)
    scored.sort(key=lambda c: c.score, reverse=True)

    current = Candidate(
        url=target.current_destination,
        title="Current workbook destination",
        destination_type=destination_type(target.current_destination),
        score=999,
    ) if target.current_destination else None

    result: List[Candidate] = []
    if current:
        result.append(current)
    for candidate in scored:
        if candidate.url and all(candidate.url != existing.url for existing in result):
            result.append(candidate)
        if len(result) >= limit:
            break
    return result


def should_review(target: ReviewTarget, include_clean_redirects: bool) -> bool:
    if target.sheet_name == "HUMAN CHECK":
        return True
    approval = target.dw_approval.upper()
    notes = target.notes.lower()
    if "REVIEW" in approval or "HUMAN" in approval:
        return True
    if any(x in notes for x in ["review", "risk", "human check", "confidence", "broad parent"]):
        return True
    if target.flags:
        return True
    return include_clean_redirects


def collect_targets(wb, include_clean_redirects: bool) -> List[ReviewTarget]:
    targets: List[ReviewTarget] = []
    for sheet_name in REVIEWABLE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = ensure_columns(ws, AI_COLUMNS)
        for row in range(2, ws.max_row + 1):
            existing_url = cell_value(ws, row, headers, "Existing URL", "Source URL")
            if not existing_url:
                continue
            current_destination = cell_value(ws, row, headers, "Destination URL", "Suggested Destination")
            dw_approval = cell_value(ws, row, headers, "DW - Approval")
            notes = cell_value(ws, row, headers, "829 Notes", "Notes", "Review Reason")
            title = cell_value(ws, row, headers, "Title 1", "Title")
            target = ReviewTarget(
                sheet_name=sheet_name,
                row_number=row,
                existing_url=existing_url,
                current_destination=current_destination,
                dw_approval=dw_approval,
                notes=notes,
                content_type=cell_value(ws, row, headers, "Content Type"),
                status_code=cell_value(ws, row, headers, "Status Code"),
                title=title,
                source_type=source_type(existing_url, title, notes),
                flags=[],
            )
            target.flags = risk_flags(target, current_destination)
            if should_review(target, include_clean_redirects):
                targets.append(target)
    return targets


def review_schema() -> Dict[str, Any]:
    return {
        "type": "json_schema",
        "name": "redirect_ai_review",
        "strict": True,
        "schema": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "decision": {"type": "string", "enum": ["approve", "review", "human_check"]},
                "risk_level": {"type": "string", "enum": ["low", "medium", "high"]},
                "recommended_destination": {"type": "string"},
                "notes_for_workbook": {"type": "string"},
                "flags": {"type": "array", "items": {"type": "string"}},
            },
            "required": ["decision", "risk_level", "recommended_destination", "notes_for_workbook", "flags"],
        },
    }


def ai_review(client: OpenAI, model: str, target: ReviewTarget, candidates: Sequence[Candidate]) -> Dict[str, Any]:
    candidate_payload = [
        {"url": c.url, "title": c.title, "destination_type": c.destination_type, "score_hint": c.score}
        for c in candidates
    ]
    prompt = {
        "task": "Review one healthcare website redirect mapping. Do not invent destinations. Prefer HUMAN CHECK when uncertain.",
        "source": {
            "url": target.existing_url,
            "title": target.title,
            "detected_type": target.source_type,
            "content_type": target.content_type,
            "status_code": target.status_code,
        },
        "current_mapping": {
            "destination_url": target.current_destination,
            "dw_approval": target.dw_approval,
            "notes": target.notes,
            "risk_flags": target.flags,
        },
        "candidate_destinations": candidate_payload,
        "rules": [
            "Manual mappings should not be overridden.",
            "Service pages should map to service destinations, not location/news/story pages.",
            "Provider pages can map to the centralized providers directory unless exact provider mapping is available.",
            "News, press releases, patient stories, and event pages require exact or very close content matches; otherwise human_check.",
            "Do not recommend external subdomains unless already supplied as the current destination and clearly appropriate.",
            "If the current destination is broad but reasonable, use review with a note rather than approve.",
        ],
    }

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "system",
                "content": "You are a cautious SEO redirect migration reviewer. Return only schema-valid JSON.",
            },
            {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
        ],
        text={"format": review_schema()},
    )

    raw = getattr(response, "output_text", "") or ""
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # Some SDK/model combinations return content blocks instead of output_text.
        as_dict = response.model_dump() if hasattr(response, "model_dump") else {}
        text_chunks: List[str] = []
        for item in as_dict.get("output", []):
            for content in item.get("content", []):
                if "text" in content:
                    text_chunks.append(content["text"])
        return json.loads("".join(text_chunks))


def apply_review(
    wb,
    target: ReviewTarget,
    review: Dict[str, Any],
    apply_ai_destination: bool,
) -> None:
    ws = wb[target.sheet_name]
    headers = ensure_columns(ws, AI_COLUMNS)
    write_cell(ws, target.row_number, headers, "AI Decision", review.get("decision", ""))
    write_cell(ws, target.row_number, headers, "AI Risk Level", review.get("risk_level", ""))
    write_cell(ws, target.row_number, headers, "AI Recommended Destination", review.get("recommended_destination", ""))
    write_cell(ws, target.row_number, headers, "AI Notes", review.get("notes_for_workbook", ""))
    write_cell(ws, target.row_number, headers, "AI Flags", ", ".join(review.get("flags", [])))

    if apply_ai_destination and review.get("decision") in {"approve", "review"}:
        recommended = norm(review.get("recommended_destination"))
        if recommended:
            write_cell(ws, target.row_number, headers, "Destination URL", recommended)


def main() -> None:
    args = parse_args()
    if not os.getenv("OPENAI_API_KEY"):
        raise SystemExit("OPENAI_API_KEY is not set. Set it locally before running this script.")

    workbook_path = args.workbook
    output_path = args.output or workbook_path.with_name(f"{workbook_path.stem}_ai_reviewed.xlsx")

    wb = load_workbook(workbook_path)
    candidates = read_candidates(wb)
    targets = collect_targets(wb, args.include_clean_redirects)
    if args.max_rows and args.max_rows > 0:
        targets = targets[: args.max_rows]

    print(f"Loaded {len(candidates):,} destination candidates.")
    print(f"Reviewing {len(targets):,} row(s).")

    client = OpenAI()
    for index, target in enumerate(targets, start=1):
        candidates_for_row = top_candidates(target, candidates)
        print(f"[{index}/{len(targets)}] {target.sheet_name} row {target.row_number}: {target.existing_url}")
        try:
            review = ai_review(client, args.model, target, candidates_for_row)
            apply_review(wb, target, review, args.apply_ai_destination)
        except Exception as exc:  # noqa: BLE001 - intentional so one bad row does not kill the run
            fallback = {
                "decision": "human_check",
                "risk_level": "high",
                "recommended_destination": target.current_destination,
                "notes_for_workbook": f"AI review failed locally: {exc}",
                "flags": ["ai_review_error"],
            }
            apply_review(wb, target, fallback, False)
        if args.sleep:
            time.sleep(args.sleep)

    wb.save(output_path)
    print(f"Saved reviewed workbook: {output_path}")


if __name__ == "__main__":
    main()
